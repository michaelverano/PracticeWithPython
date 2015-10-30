#!	python3
#	multiplicationTable.py - takes the second argument command line when running python and 
#	creates a multiplication table on Excel. The second argument must be an int.

import sys
import openpyxl
from openpyxl.cell import get_column_letter

import logging
logging.basicConfig(level=logging.DEBUG, format=' %(asctime)s %(levelname)s - %(message)s')

def multiplicationTable(multiplier):
	logging.debug('multiplication table initialized...')

	# Create the workbook
	wb = openpyxl.Workbook()
	sheet = wb.get_active_sheet()
	logging.debug('Workbook created and active sheet initialized...')

	# Create the first rows
	for rows in range(1, multiplier + 1):
		sheet[get_column_letter(1) + str(rows)] = rows
	logging.debug('First row created...')

	# Create the first columns
	for columns in range(1, multiplier + 1):
		sheet[get_column_letter(columns) + str(1)] = columns
	logging.debug('First column created...')

	# create the remaining rows with a formula
	for rows in range(2, multiplier + 1):
		# Create columns.
		for columns in range(2, multiplier + 1):
			# formula for multiplying the tables.
			formula = '=A' + str(rows) + '*' + get_column_letter(columns) + str(1)  
			sheet[get_column_letter(columns) + str(rows)] = formula
	logging.debug('Multiplication table made...')

	# Save into a worksheet.
	wb.save('multiplicationTable.xlsx')
	print('\n\nMultiplication table saved.')

if __name__ == '__main__':
	
	if len(sys.argv) == 3 and sys.argv[2] == 'debug':
		print('Debug mode on...')
		logging.debug('number of arguments is... ' + str(len(sys.argv)))
	else:
		logging.disable(logging.CRITICAL)

	# Run the program.
	if len(sys.argv) == 1:
		multiplier = input('What number do you want to multiply? : ')
		multiplicationTable(int(multiplier))
	else:
		multiplicationTable(int(sys.argv[1]))
		